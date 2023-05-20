from PySimpleGUI import PySimpleGUI as sg
import openpyxl

book = openpyxl.Workbook()
# criação da planilha

book.create_sheet('Cadastro')

cadastro_page = book['Cadastro']
cadastro_page.append(['Nome', 'Sexo', 'Endereço', 'Cidade'])

# Layout
sg.theme('Reddit')
layout = [
    [sg.Text('Cadastro de Pacientes')],
    [sg.Text('Nome Completo')],
    [sg.Input(key='nome', size=(40, 1))],
    [sg.Text('Sexo')],
    [sg.Input(key='sexo', size=(4, 1))],
    [sg.Text('Endereço')],
    [sg.Input(key='endereco', size=(40, 1))],
    [sg.Text('Cidade')],
    [sg.Input(key='cidade', size=(40, 1))],
    [sg.Button('Cadastrar'), sg.Button('Sair')],
    [sg.Text(key='message')]
]

# Janela
window = sg.Window('Cadastro de Pacientes', layout)

while True:
    eventos, valores = window.read()
    if eventos == sg.WINDOW_CLOSED or eventos == 'Sair':
        break
# Criação das variaveis que irão receber os valores da janela
    if eventos == 'Cadastrar':
        input_nome = valores['nome']
        input_sexo = valores['sexo']
        input_endereco = valores['endereco']
        input_cidade = valores['cidade']
        cadastro_page.append(
            [input_nome, input_sexo, input_endereco, input_cidade])
    book.save('Planilha de Pacientes.xlsx')
    window['message'].update("Paciente " + input_nome +
                             " salvo com sucesso!")


print(book.sheetnames)
window.close()
